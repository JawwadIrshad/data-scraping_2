import pandas as pd

main_file = "n8n good data - Sheet1.csv"
match_file = "Untitled spreadsheet - Sheet2.csv"
output_file = "highlighted.xlsx"

# Read both files
main_df = pd.read_csv(main_file)
match_df = pd.read_csv(match_file)

# Normalize names for matching
match_names = set(match_df.iloc[:, 0].str.strip().str.lower())

# Add status column
main_df["Status"] = main_df.iloc[:, 0].str.strip().str.lower().isin(match_names)

# Sort: TRUE on top
main_df = main_df.sort_values(by="Status", ascending=False).reset_index(drop=True)

# Save to Excel
main_df.to_excel(output_file, index=False)

# Highlight TRUE rows
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook(output_file)
ws = wb.active
highlight = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

status_col_index = main_df.columns.get_loc("Status") + 1
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=status_col_index).value:
        for col in range(1, len(main_df.columns) + 1):
            ws.cell(row=row, column=col).fill = highlight

wb.save(output_file)
print(f"✅ Highlighted Excel file saved as '{output_file}'")
